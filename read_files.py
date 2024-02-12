import openpyxl
from pathlib import Path
import os
import logging
import tempfile
import shutil
import re



from win32com import client
from execute_pdf import merge_excel_sheet_to_pdf
win32_excel = client.Dispatch("Excel.Application")




# function to create a temp excel file and save it in temp directory
def create_excel_in_temp(excel_file):
    temp_excelfile_name = tempfile.NamedTemporaryFile(suffix=excel_file.filename,prefix="", delete=False)
    excel_file.save(temp_excelfile_name.name)
    return temp_excelfile_name




class WriteToExcel:
    def __init__(self, file_name, sheet_name, start_depth, end_depth, column_start_range, column_end_range, student_id_column = "B") -> None:
        self.file_name = file_name
        # create a excel file in temp location and save the sheet in that location
        self.temp_excelfile_name = create_excel_in_temp(self.file_name)
        self.sheet_name = sheet_name
        self.start_depth = start_depth
        self.end_depth = end_depth
        self.column_start_range = column_start_range
        self.column_end_range = column_end_range
        self.iterating_range = [i for i in range(self.start_depth, self.end_depth+1)]
        self.student_id_column = student_id_column
        self.column_array_range = self.create_column_array()
        self.work_book = openpyxl.load_workbook(self.temp_excelfile_name.name)
        self.sheet = self.work_book[self.sheet_name]
        self.sheet_marks_position = self.create_sheet_marks_position()
        self.original_length_of_iterating_range = len(self.iterating_range)


    
    def range_char(self,start, stop):
        return (chr(n) for n in range(ord(start), ord(stop) + 1))

    def create_column_array(self):
        my_array = []
        for char in self.range_char(self.column_start_range, self.column_end_range):
            my_array.append(char)
        return my_array
    
    def create_sheet_marks_position(self):
        my_dict = {}
        
        for letter in self.column_array_range:
            value = self.sheet[f"{letter}{self.start_depth-1}"].value.split("(")[0].strip()
            my_dict[value] = f"{letter}"

        return my_dict
    
    def find_student_row_number(self, student_id):
        row_no = 0
        for i in self.iterating_range:
            if student_id == self.sheet[f"{self.student_id_column}{i}"].value:
                row_no = i
                self.iterating_range.remove(i)
                break
        completion_percentage = int(((self.original_length_of_iterating_range -len(self.iterating_range))/self.original_length_of_iterating_range)*100)
        return row_no, completion_percentage



class ReadFromExcel:
    def __init__(self, folder_path, marks_dict, sheet_name, excel_object,merge_excel,sse,isGroupCourseWork):
        self.folder_path = folder_path 
        self.marks_dict = marks_dict
        self.sheet_name = sheet_name
        self.excel_object = excel_object
        self.merge_excel = merge_excel
        self.sse = sse
        self.is_group_coursework = isGroupCourseWork
        self.desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')


        # create a logger
        # desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        logging.basicConfig(encoding ='utf-8',level=logging.ERROR,format='%(levelname)s:%(message)s')
        self.logger = logging.getLogger(__name__)

        ch = logging.FileHandler(filename = f'{self.desktop}/{Path(self.folder_path).stem}.log',mode='w')
        ch.setLevel(logging.ERROR)
        formatter = logging.Formatter('%(levelname)s - %(message)s')
        ch.setFormatter(formatter)
        self.logger.addHandler(ch)
       
    def read_folder(self):
        path = Path(self.folder_path)
        """
        Reading the each folder of student and taking the student id
        """
        for p in path.iterdir():
            self.sse.publish({"name":f"{' '.join(p.stem.split()[1:])}"},type='showInfo')
            if self.is_group_coursework:
                try:
                    student_id = int(os.path.basename(p).split()[0])
                except:
                    self.logger.error(f"Folder of student {os.path.basename(p).split()[0]} does not contain student id")
                    continue
           
            self.read_from_excel_file(p)
        # saving the rte provided excel sheet in the desktop location
        self.excel_object.temp_excelfile_name.close()
        shutil.move(self.excel_object.temp_excelfile_name.name,os.path.join(self.desktop,self.excel_object.file_name.filename))

            
    def read_from_excel_file(self,p):
        excel_file = None

        student_name = " ".join(p.stem.split()[1:]) # extracting student name form the folder
        for file in p.rglob("*.xlsx"):
            if re.match(r"\d+\s+[a-z A-Z. ()]+.xlsx",file.name):
                wb = openpyxl.load_workbook(file, data_only=True,read_only=True)
                ws = wb[self.sheet_name]
                # from helper_function import generate_marks_dictonary

                # generate_marks_dictonary(self.excel_object.sheet,ws,self.excel_object.start_depth, self.excel_object.column_start_range, self.excel_object.column_end_range)
                


                student_id = int(file.stem.split(" ")[0])
                student_name = " ".join(file.stem.split(" ")[1:])
                excel_file = file
                wb = openpyxl.load_workbook(file, data_only=True,read_only=True)
                ws = wb[self.sheet_name]
                marks_value_dict = {}
                for key, value in self.marks_dict.items():
                    # has used this try block to convert 0 string into integer
                    try:
                        marks_value_dict[key] = float(ws[value].value)
                    except Exception:
                        marks_value_dict[key] = 0
                  

                self.write_to_main_docs(student_id, marks_value_dict,student_name)
                wb.close()
                if not self.is_group_coursework:
                    break
        if self.merge_excel:
            for file in p.rglob("*.pdf"):
                merge_excel_sheet_to_pdf(str(p),excel_file=excel_file,pdf_file=file)
                break
                
    def write_to_main_docs(self,student_id, marks_value_dict,student_name):
 
        if len(marks_value_dict.keys()) or len(self.excel_object.sheet_marks_position.keys()):            
            student_row_num,completion_percentage = self.excel_object.find_student_row_number(student_id)
            # this code is to broadcast the percentage
            self.sse.publish({"percentage":completion_percentage},type='showPercent')
            if student_row_num >0:
                for key, value in marks_value_dict.items():
                    if key in self.excel_object.sheet_marks_position:
                        self.excel_object.sheet[f"{self.excel_object.sheet_marks_position[key]}{student_row_num}"] = value
                self.excel_object.work_book.save(self.excel_object.temp_excelfile_name.name)
            else:
                self.logger.error(f"Error in london met id of student {student_name} having id {student_id}")
        
        else:
            print(f"Marks dictonary is not  valid one")
                
                


