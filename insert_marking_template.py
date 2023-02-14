from pathlib import Path
import shutil
import openpyxl

def iterate_folder(folder_path, excel_sheet,cell_infos,sheet_name):
    path = Path(folder_path)

    for each in path.iterdir():
        if not each.is_file():
            copy_excel_sheet(excel_sheet,each,cell_infos,sheet_name)
        else:
            folder_name = create_folder(each)
            copy_excel_sheet(excel_sheet,folder_name,cell_infos,sheet_name)

        
def copy_excel_sheet(excel_sheet,folder,cell_infos,sheet_name="Grading Sheet"):
    excel_file_name = Path(excel_sheet).name
    file_extension = Path(excel_sheet).suffix
    student_folder_info= Path(folder).name

    # extracting student name and student id

    student_infos = student_folder_info.split( )
    student_id = student_infos[0]
    student_name = " ".join(student_infos[1:])

    shutil.copy(excel_sheet,folder)
    shutil.move(f"{folder}/{excel_file_name}",f"{folder}/{student_folder_info}{file_extension}")
    excelFilePath = f"{folder}/{student_folder_info}{file_extension}"


    """
    code to insert student name and student id inside excel sheet file
    """
    wb = openpyxl.load_workbook(excelFilePath,data_only=True)
    ws = wb[sheet_name]
    for key,value in cell_infos.items():
        if key =="name":
            ws[value] = student_name
        else:
            ws[value] = student_id
    wb.close()
    wb.save(excelFilePath)
        

def create_folder(file):
    folder_location = Path(file).parent.absolute()
    student_info = Path(file).stem
    print(f"student_info: {student_info}")

    new_folder_name = Path(f"{folder_location}/{student_info}")
    new_folder_name.mkdir(parents=True, exist_ok=True)
    shutil.move(file,new_folder_name)
    return new_folder_name



    


        
       